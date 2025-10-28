"""
Excel Scanner Utilities - Updated with GCS Integration
Các function để scan và phân tích Excel files với GCS storage
"""

import pandas as pd
from datetime import datetime
from pathlib import Path
import json
from openpyxl import load_workbook
import xlrd
import re
import os
import numpy as np
from difflib import get_close_matches
import tempfile
import shutil
from typing import Optional, List, Dict, Any
from utils.gcs_storage import GCSFileManager, clean_for_json

# ==== Constants ====
ANCHORS = [
    "COD", "COD. E", "DEBITORE", "TIPOLOGIA", "INDIRIZZO", "CITTA", "CITTÃ€",
    "CAP", "LUOGO", "DATA", "FIRMATARI", "NDG", "CAPITALE", "INTERE", "GBV",
    "CODICE FISCALE", "P.IVA", "CF", "CITTADINANZA", "REGIONE", "NOMINATIVO",
]

# ==== Helper Functions ====
def _is_empty_like(x) -> bool:
    """Kiểm tra xem giá trị có rỗng/vô nghĩa không"""
    if x is None: 
        return True
    if isinstance(x, float) and pd.isna(x): 
        return True
    s = str(x).strip()
    return s == "" or s.lower() == "nan" or s.startswith("Unnamed:")

def count_unnamed_columns(columns) -> int:
    """Đếm số cột Unnamed hoặc rỗng"""
    count = 0
    for col in columns:
        col_str = str(col).strip()
        if col_str.startswith("Unnamed:") or col_str == "" or col_str.lower() == "nan":
            count += 1
    return count

def count_anchor_matches(columns, anchors=ANCHORS, cutoff=0.75) -> int:
    """Đếm số anchor keywords được tìm thấy"""
    cols_upper = [str(c).strip().upper() for c in columns if not _is_empty_like(c)]
    matches = 0
    for anchor in anchors:
        if get_close_matches(anchor.upper(), cols_upper, n=1, cutoff=cutoff):
            matches += 1
    return matches

# ==== Core Functions ====
def find_best_header_row(file_path: str, sheet_name: str, max_scan_rows: int = 20) -> Optional[Dict]:
    """
    Quét N dòng đầu và chọn dòng có:
    1. Ít Unnamed columns nhất
    2. Nếu bằng nhau thì chọn dòng có nhiều anchor matches nhất
    3. Nếu vẫn bằng nhau thì chọn dòng đầu tiên
    """
    candidates = []
    
    print(f"      🔍 Quét {max_scan_rows} dòng đầu để tìm header tốt nhất...")
    
    for row_idx in range(max_scan_rows):
        try:
            # Đọc chỉ columns (nrows=0)
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=row_idx, nrows=0)
            columns = [str(col).strip() for col in df.columns.tolist()]
            
            if len(columns) == 0:  # Skip empty rows
                continue
            
            unnamed_count = count_unnamed_columns(columns)
            anchor_matches = count_anchor_matches(columns)
            valid_cols = len(columns) - unnamed_count
            
            candidates.append({
                'row_idx': row_idx,
                'columns': columns,
                'total_cols': len(columns),
                'unnamed_count': unnamed_count,
                'valid_cols': valid_cols,
                'anchor_matches': anchor_matches,
                'unnamed_ratio': unnamed_count / len(columns) if len(columns) > 0 else 1.0
            })
            
            print(f"        Dòng {row_idx+1}: {len(columns)} cột, {unnamed_count} Unnamed ({unnamed_count/len(columns)*100:.0f}%), {anchor_matches} anchors")
            
        except Exception as e:
            print(f"        Dòng {row_idx+1}: ❌ Lỗi đọc - {str(e)[:50]}")
            continue
    
    if not candidates:
        print("        ❌ Không tìm thấy dòng hợp lệ nào")
        return None
    
    # Sắp xếp theo: 
    # 1. Ít Unnamed nhất (tăng dần)
    # 2. Nhiều anchor matches nhất (giảm dần) 
    # 3. Dòng đầu tiên (tăng dần)
    candidates.sort(key=lambda x: (x['unnamed_count'], -x['anchor_matches'], x['row_idx']))
    
    best = candidates[0]
    print(f"        ✅ Chọn dòng {best['row_idx']+1}: {best['valid_cols']}/{best['total_cols']} cột hợp lệ, {best['anchor_matches']} anchors")
    
    return best

def get_sample_data(file_path: str, sheet_name: str, header_row_idx: int) -> Optional[List[str]]:
    """Lấy 1 dòng data mẫu sau header"""
    try:
        # Đọc 1 dòng data sau header
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_idx, nrows=1)
        if len(df) > 0:
            # Chuyển về string và xử lý NaN
            sample_row = []
            for val in df.iloc[0]:
                if pd.isna(val):
                    sample_row.append("")
                else:
                    sample_row.append(str(val))
            return sample_row
        return None
    except Exception as e:
        print(f"        ⚠️ Không lấy được sample data: {e}")
        return None

def save_processed_data_to_gcs(file_path: str, sheet_name: str, header_row_idx: int, 
                            original_filename: str, gcs_manager: GCSFileManager) -> Optional[Dict]:
    """Lưu processed data lên GCS"""
    try:
        df_cleaned = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_idx)
        
        processed_info = gcs_manager.save_processed_file(
            df=df_cleaned,
            original_filename=original_filename,
            sheet_name=sheet_name,
            header_row_idx=header_row_idx
        )
        
        print(f"        💾 Đã lưu processed data lên GCS")
        return processed_info
        
    except Exception as e:
        print(f"        ❌ Lỗi khi lưu processed data: {e}")
        return None

# ==== Main Scanning Functions ====
def scan_single_excel_file(file_path: str, max_scan_rows: int = 10, 
                        gcs_manager: Optional[GCSFileManager] = None,
                        original_filename: Optional[str] = None) -> List[Dict]:
    """
    Scan một file Excel và trả về thông tin các sheet
    Có thể lưu processed files lên GCS nếu có gcs_manager
    """
    results = []
    
    try:
        # Xác định loại file
        file_ext = Path(file_path).suffix.lower()
        
        if file_ext == ".xlsx":
            # Sử dụng openpyxl
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        elif file_ext == ".xls":
            # Sử dụng xlrd
            wb = xlrd.open_workbook(file_path)
            sheet_names = wb.sheet_names()
        else:
            raise ValueError(f"Unsupported file format: {file_ext}")
        
        # Scan từng sheet
        for i, sheet_name in enumerate(sheet_names):
            print(f"   📊 Phân tích sheet [{i+1}/{len(sheet_names)}]: {sheet_name}")
            
            best_header = find_best_header_row(file_path, sheet_name, max_scan_rows)
            
            if best_header and best_header['valid_cols'] > 0:
                # Có header hợp lệ
                sample_data = get_sample_data(file_path, sheet_name, best_header['row_idx'])
                
                # Lưu processed data lên GCS nếu có gcs_manager
                processed_info = None
                if gcs_manager and original_filename:
                    processed_info = save_processed_data_to_gcs(
                        file_path, sheet_name, best_header['row_idx'], 
                        original_filename, gcs_manager
                    )
                
                results.append({
                    "sheet_name": sheet_name,
                    "have_header": True,
                    "header_row_idx": best_header['row_idx'],  # Index bắt đầu từ 0
                    "header_at_row": best_header['row_idx'] + 1,  # Số dòng thực tế (bắt đầu từ 1)
                    "columns": best_header['columns'],
                    "sample_data": sample_data,
                    "header_quality": {
                        "total_cols": best_header['total_cols'],
                        "valid_cols": best_header['valid_cols'],
                        "unnamed_count": best_header['unnamed_count'],
                        "unnamed_ratio": best_header['unnamed_ratio'],
                        "anchor_matches": best_header['anchor_matches']
                    },
                    "processed_file_info": processed_info
                })
            else:
                # Không tìm thấy header hợp lệ
                results.append({
                    "sheet_name": sheet_name,
                    "have_header": False,
                    "header_row_idx": None,
                    "header_at_row": None,  # Không có header
                    "columns": [],
                    "sample_data": None,
                    "header_quality": None,
                    "processed_file_info": None
                })
        
    except Exception as e:
        print(f"❌ Lỗi đọc file {file_path}: {e}")
        raise
    
    return results

def scan_uploaded_file_with_gcs(file_content: bytes, filename: str, 
                            max_scan_rows: int = 20, 
                            save_to_gcs: bool = True,
                            gcs_bucket_name: str = None) -> Dict[str, Any]:
    """
    Scan file Excel được upload (từ bytes) và lưu lên GCS
    """
    # Khởi tạo GCS manager nếu cần
    gcs_manager = None
    if save_to_gcs:
        try:
            gcs_manager = GCSFileManager(gcs_bucket_name) if gcs_bucket_name else GCSFileManager()
        except Exception as gcs_error:
            print(f"⚠️ Không thể khởi tạo GCS manager: {gcs_error}")
            # Continue without GCS
            save_to_gcs = False
    
    # Tạo temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix=Path(filename).suffix) as temp_file:
        temp_file.write(file_content)
        temp_file_path = temp_file.name
    
    try:
        print(f"🔍 Scan uploaded file: {filename}")
        
        # Scan file
        sheet_info = scan_single_excel_file(
            temp_file_path, 
            max_scan_rows, 
            gcs_manager if save_to_gcs else None,
            filename 
        )
        
        # Tạo scan result
        scan_result = {
            "filename": filename,
            "file_size": len(file_content),
            "scan_timestamp": datetime.now().isoformat(),
            "sheets": sheet_info,
            "total_sheets": len(sheet_info),
            "sheets_with_header": sum(1 for sheet in sheet_info if sheet.get('have_header', False))
        }
        
        # Clean data for JSON
        scan_result = clean_for_json(scan_result)
        
        # Lưu complete result lên GCS nếu cần
        gcs_storage_info = None
        if save_to_gcs and gcs_manager:
            try:
                # Tập hợp processed files info
                processed_files = [
                    sheet.get('processed_file_info') 
                    for sheet in sheet_info 
                    if sheet.get('processed_file_info')
                ]
                
                gcs_storage_info = gcs_manager.save_complete_scan_result(
                    file_content=file_content,
                    original_filename=filename,
                    scan_result=scan_result,
                    processed_files_info=processed_files
                )
                
                print(f"💾 Complete scan result saved to GCS")
                
            except Exception as e:
                print(f"⚠️ Lỗi lưu lên GCS: {e}")
                gcs_storage_info = {"error": str(e)}
        
        # Thêm GCS info vào result
        scan_result["gcs_storage"] = gcs_storage_info
        
        return scan_result
        
    finally:
        # Xóa temp file
        try:
            os.unlink(temp_file_path)
        except Exception as e:
            print(f"⚠️ Không thể xóa temp file: {e}")

def scan_multiple_uploaded_files_with_gcs(files_data: List[tuple], 
                                        max_scan_rows: int = 20,
                                        save_to_gcs: bool = True,
                                        gcs_bucket_name: str = None) -> List[Dict]:
    """
    Scan nhiều file Excel và lưu lên GCS
    
    Args:
        files_data: List of (file_content, filename) tuples
        max_scan_rows: Số dòng để scan header
        save_to_gcs: Có lưu lên GCS không
        gcs_bucket_name: Tên GCS bucket (optional)
    
    Returns:
        List kết quả scan từng file
    """
    results = []
    
    for i, (file_content, filename) in enumerate(files_data):
        try:
            print(f"🔍 Scan file [{i+1}/{len(files_data)}]: {filename}")
            
            result = scan_uploaded_file_with_gcs(
                file_content=file_content,
                filename=filename,
                max_scan_rows=max_scan_rows,
                save_to_gcs=save_to_gcs,
                gcs_bucket_name=gcs_bucket_name
            )
            
            results.append({
                "filename": filename,
                "success": True,
                "result": result
            })
            
        except Exception as e:
            print(f"❌ Lỗi scan file {filename}: {e}")
            results.append({
                "filename": filename,
                "success": False,
                "error": str(e)
            })
    
    return results

# ==== Legacy Functions for Backward Compatibility ====
def scan_uploaded_file(file_content: bytes, filename: str, max_scan_rows: int = 20) -> Dict[str, Any]:
    """
    Legacy function - scan file without GCS (for backward compatibility)
    """
    return scan_uploaded_file_with_gcs(
        file_content=file_content,
        filename=filename,
        max_scan_rows=max_scan_rows,
        save_to_gcs=False
    )

# ==== Utility Functions ====
def print_scan_results(results: List[Dict]):
    """In kết quả quét một cách dễ đọc"""
    for file_info in results:
        print(f"\n🔍 File: {Path(file_info['file']).name}")
        print("=" * 60)
        
        for sheet in file_info['sheets']:
            print(f"\n📊 Sheet: {sheet['sheet_name']}")
            
            if sheet['have_header']:
                quality = sheet['header_quality']
                print(f"   ✅ Header tìm thấy tại dòng {sheet['header_at_row']}")
                print(f"   📋 Chất lượng: {quality['valid_cols']}/{quality['total_cols']} cột hợp lệ ({quality['unnamed_ratio']*100:.0f}% Unnamed)")
                print(f"   🎯 Anchor matches: {quality['anchor_matches']}")
                
                if len(sheet['columns']) <= 6:
                    print(f"   📝 Columns: {sheet['columns']}")
                else:
                    print(f"   📝 Columns ({len(sheet['columns'])}): {sheet['columns'][:3]} ... {sheet['columns'][-2:]}")
                
                if sheet.get('processed_file_info'):
                    processed = sheet['processed_file_info']
                    print(f"   🧹 Processed: {processed['data_rows']} rows × {processed['data_cols']} cols")
                    print(f"   📄 GCS URI: {processed['gcs_uri']}")
            else:
                print("   ❌ Không tìm thấy header hợp lệ")

def debug_header_detection(file_path: str, sheet_name: Optional[str] = None, max_scan_rows: int = 30):
    """Debug chi tiết quá trình phát hiện header"""
    print(f"🔍 Debug header detection: {file_path}")
    
    try:
        if file_path.endswith('.xlsx'):
            wb = load_workbook(filename=file_path, read_only=True)
            sheets_to_check = [sheet_name] if sheet_name else wb.sheetnames
            wb.close()
        else:
            wb = xlrd.open_workbook(file_path)
            sheets_to_check = [sheet_name] if sheet_name else wb.sheet_names()
        
        for sname in sheets_to_check:
            print(f"\n📊 Sheet: {sname}")
            best_header = find_best_header_row(file_path, sname, max_scan_rows)
            
            if best_header:
                print(f"🎯 BEST HEADER: Dòng {best_header['row_idx']+1}")
                print(f"   Columns: {best_header['columns'][:8]}{'...' if len(best_header['columns']) > 8 else ''}")
            else:
                print("❌ Không tìm thấy header phù hợp")
            
    except Exception as e:
        print(f"❌ Lỗi: {e}")

def save_scan_results_to_json(results: List[Dict], output_folder: str = "scanned_schema") -> str:
    """Lưu kết quả scan ra file JSON (local fallback)"""
    runtime = datetime.now()
    output_path = Path(output_folder) / f"scan_results_{runtime.strftime('%Y%m%d_%H%M%S')}.json"
    output_path.parent.mkdir(exist_ok=True)
    
    # Làm sạch data trước khi lưu JSON
    clean_results = clean_for_json(results)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(clean_results, f, ensure_ascii=False, indent=4)
    
    return str(output_path)